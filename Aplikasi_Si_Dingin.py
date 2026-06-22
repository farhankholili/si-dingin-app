import streamlit as st
import pandas as pd
import os
import io  # Untuk memproses file Excel di memori RAM
import openpyxl
import random
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as ExcelImage

# Ambil model database baru yang sudah Multi-Tenant
from database import db, Order, DetailOrder, User, Layanan, Bengkel
import os

# 1. Pastikan database dan tabel sudah terbuat sempurna
db.connect(reuse_if_open=True)
db.create_tables([Bengkel, User, Layanan, Order, DetailOrder])

# 2. SUNTIK DATA AWAL (BIAR GAK GAGAL BUAT AKUN)
def suntik_data_awal():
    # Buat Bengkel Pertama (jika belum ada)
    bengkel_utama, created = Bengkel.get_or_create(
        nama_bengkel="Mutiara AC Center",
        defaults={
            "alamat": "Situbondo",
            "no_telp": "08123456789",
            "owner_name": "Farhan Kholili"
        }
    )
    
    # Buat Akun Super Admin bawaan untuk login pertama (jika belum ada)
    # Catatan: Di kodingan asli Bos, pastikan password-nya di-hash kalau pakai sistem keamanan hash!
    User.get_or_create(
        username="admin",
        defaults={
            "bengkel": bengkel_utama,
            "password": "admin", # Silakan ganti sesuai kebutuhan Bos
            "nama_lengkap": "Super Admin Farhan",
            "role": "super_admin",
            "is_active": True
        }
    )

# Jalankan fungsi suntik data
suntik_data_awal()

# Buat ulang database baru gres dengan struktur Multi-Tenant yang sempurna
db.connect(reuse_if_open=True)
db.create_tables([Bengkel, User, Layanan, Order, DetailOrder])

# Konfigurasi Halaman Utama
st.set_page_config(page_title="Si Dingin - Sistem Integrasi SaaS", layout="wide")

# Hubungkan Database
db.connect(reuse_if_open=True)
db.create_tables([Bengkel, User, Layanan, Order, DetailOrder])

# --------------------------------------------------
# INISIALISASI SESSION STATE (PAPAN SATPAM)
# --------------------------------------------------
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
if "user_role" not in st.session_state:
    st.session_state.user_role = None
if "user_fullname" not in st.session_state:
    st.session_state.user_fullname = ""
if "user_id" not in st.session_state:
    st.session_state.user_id = None
if "username" not in st.session_state:
    st.session_state.username = ""

# Menyimpan Data Ruang Kamar Bengkel Tenant
if "bengkel_id" not in st.session_state:
    st.session_state.bengkel_id = None
if "nama_usaha" not in st.session_state:
    st.session_state.nama_usaha = "NOTA DIGITAL JASA REPARASI & SERVICE AC"
if "alamat_usaha" not in st.session_state:
    st.session_state.alamat_usaha = "AC Dingin, Kerja Transparan, Keluarga Nyaman"
if "kontak_usaha" not in st.session_state:
    st.session_state.kontak_usaha = "0812-XXXX-XXXX"

if "keranjang_kerja" not in st.session_state:
    st.session_state.keranjang_kerja = []
if "nota_siap_download" not in st.session_state:
    st.session_state.nota_siap_download = None

# Fitur Lupa Password
if "forgot_mode" not in st.session_state:
    st.session_state.forgot_mode = False
if "otp_terkirim" not in st.session_state:
    st.session_state.otp_terkirim = None
if "otp_terverifikasi" not in st.session_state:
    st.session_state.otp_terverifikasi = False

# ==================================================
# FUNGSI ENGINERING: KIRIM OTP EMAIL
# ==================================================
def kirim_email_otp(email_tujuan, otp_code):
    email_pengirim = "owner.sidingin@gmail.com" 
    password_aplikasi = "xxxx xxxx xxxx xxxx" # ⚠️ Masukkan 16 Digit App Password Google Bos

    msg = MIMEMultipart()
    msg['From'] = email_pengirim
    msg['To'] = email_tujuan
    msg['Subject'] = "🔐 KODE OTP: Pemulihan Sistem Si Dingin"

    body = f"""
    Halo Bos Farhan,
    
    Sistem mendeteksi permintaan pemulihan akun untuk dashboard Owner Si Dingin.
    Berikut adalah kode verifikasi OTP Bos:
    
    👉 {otp_code} 👈
    
    Masukkan kode di atas ke dalam aplikasi untuk membuka akses perubahan akun baru.
    
    Salam Sukses,
    Sistem Integrasi Si Dingin ❄️
    """
    msg.attach(MIMEText(body, 'plain'))
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(email_pengirim, password_aplikasi)
        server.sendmail(email_pengirim, email_tujuan, msg.as_string())
        server.quit()
        return True
    except Exception:
        return False

# ==================================================
# FUNGSI SAKTI: CETAK NOTA PREMIUM
# ==================================================
def buat_file_excel_nota(order_obj, list_detail_pekerjaan):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nota Digital"
    ws.views.sheetView[0].showGridLines = True
    
    font_judul = Font(name="Arial", size=16, bold=True, color="1E3A8A")
    font_sub_judul = Font(name="Arial", size=9, italic=True, color="555555")
    font_section = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_header_tabel = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=10, bold=True)
    font_data = Font(name="Arial", size=10)
    font_garansi = Font(name="Arial", size=9, italic=True)
    
    fill_biru_tua = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    fill_biru_tabel = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='B0B0B0'), right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'), bottom=Side(style='thin', color='B0B0B0')
    )
    
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 15
    
    if os.path.exists("logo.png"):
        try:
            img = ExcelImage("logo.png")
            img.width = 80
            img.height = 80
            ws.add_image(img, 'A1')
        except Exception:
            pass
        
    ws.merge_cells("B2:F2")
    ws["B2"] = st.session_state.nama_usaha.upper()
    ws["B2"].font = font_judul
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws.merge_cells("B3:F3")
    ws["B3"] = f"{st.session_state.alamat_usaha} | Telp: {st.session_state.kontak_usaha}"
    ws["B3"].font = font_sub_judul
    ws["B3"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws.merge_cells("A6:F6")
    ws["A6"] = " DATA NOTA & PELANGGAN"
    ws["A6"].font = font_section
    ws["A6"].fill = fill_biru_tua
    ws["A6"].alignment = Alignment(vertical="center")
    ws.row_dimensions[6].height = 22
    
    nama_tek = order_obj.teknisi.nama_lengkap if order_obj.teknisi else "Tanpa Teknisi"
    hp_tek = "-"
    if order_obj.teknisi:
        hp_tek = getattr(order_obj.teknisi, 'no_hp', getattr(order_obj.teknisi, 'no_telp', '-'))
    
    status_db = getattr(order_obj, "status_pembayaran", "Belum Dibayar")
    status_text = "LUNAS" if status_db == "Lunas" else "BELUM LUNAS"
    
    nomor_telepon_pelanggan = "-"
    if hasattr(order_obj, 'no_telp_pelanggan'): nomor_telepon_pelanggan = order_obj.no_telp_pelanggan
    elif hasattr(order_obj, 'no_telp'): nomor_telepon_pelanggan = order_obj.no_telp
    elif hasattr(order_obj, 'no_hp'): nomor_telepon_pelanggan = order_obj.no_hp
        
    ws["A7"] = "No. Nota / Invoice"; ws["B7"] = order_obj.no_invoice
    ws["A8"] = "Tanggal Pekerjaan"; ws["B8"] = str(order_obj.tanggal_kerja)
    ws["A9"] = "Teknisi Lapangan"; ws["B9"] = f"{nama_tek} ({hp_tek})"
    ws["A10"] = "Status Pembayaran"; ws["B10"] = status_text
    
    alamat_p_text = getattr(order_obj, 'alamat_pelanggan', '-')
        
    ws["D7"] = "Nama Pelanggan"; ws["E7"] = order_obj.nama_pelanggan
    ws["D8"] = "No. Telepon / WA"; ws["E8"] = nomor_telepon_pelanggan
    ws["D9"] = "Alamat Pelanggan"; ws["E9"] = alamat_p_text
    ws["D10"] = "Metode Pembayaran"; ws["E10"] = "Cash / Transfer"
    
    for r in range(7, 11):
        ws.row_dimensions[r].height = 18
        ws[f"A{r}"].font = font_bold; ws[f"B{r}"].font = font_data
        ws[f"D{r}"].font = font_bold; ws[f"E{r}"].font = font_data
    
    ws.merge_cells("A12:F12")
    ws["A12"] = " DETAIL PEKERJAAN, SPAREPART & BIAYA"
    ws["A12"].font = font_section
    ws["A12"].fill = fill_biru_tua
    ws["A12"].alignment = Alignment(vertical="center")
    ws.row_dimensions[12].height = 22
    
    headers = ["No", "Deskripsi Pekerjaan / Sparepart", "Qty (Unit/Pcs)", "Harga Satuan", "Total Jasa & Bahan"]
    ws.cell(row=13, column=1, value=headers[0])
    ws.cell(row=13, column=2, value=headers[1])
    ws.cell(row=13, column=3, value=headers[2])
    ws.merge_cells("C13:D13")
    ws.cell(row=13, column=5, value=headers[3])
    ws.cell(row=13, column=6, value=headers[4])
    
    ws.row_dimensions[13].height = 25
    for col_idx in [1, 2, 3, 5, 6]:
        c = ws.cell(row=13, column=col_idx)
        c.font = font_header_tabel; c.fill = fill_biru_tabel
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border_thin
    
    current_row = 14
    total_akhir = 0
    for idx, item in enumerate(list_detail_pekerjaan, start=1):
        ws.row_dimensions[current_row].height = 20
        sub_item = item["harga"] * item["jumlah"]
        total_akhir += sub_item
        
        ws.cell(row=current_row, column=1, value=idx).alignment = Alignment(horizontal="center", vertical="center")
        teks_ruangan = f" ({item['lokasi_ruang']})" if item['lokasi_ruang'] and item['lokasi_ruang'] != "-" else ""
        ws.cell(row=current_row, column=2, value=f"{item['nama_layanan']}{teks_ruangan}").alignment = Alignment(vertical="center")
        
        ws.cell(row=current_row, column=3, value=item["jumlah"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
        
        c_harga = ws.cell(row=current_row, column=5, value=item["harga"])
        c_harga.number_format = '"Rp "#,##0'; c_harga.alignment = Alignment(horizontal="right", vertical="center")
        
        c_sub = ws.cell(row=current_row, column=6, value=sub_item)
        c_sub.number_format = '"Rp "#,##0'; c_sub.alignment = Alignment(horizontal="right", vertical="center")
        
        for col_idx in range(1, 7):
            ws.cell(row=current_row, column=col_idx).font = font_data
            ws.cell(row=current_row, column=col_idx).border = border_thin
        current_row += 1
    
    ws.row_dimensions[current_row].height = 22
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    c_tot_label = ws.cell(row=current_row, column=1, value="TOTAL PEMBAYARAN :")
    c_tot_label.font = font_bold; c_tot_label.alignment = Alignment(horizontal="right", vertical="center")
    
    c_tot_val = ws.cell(row=current_row, column=6, value=total_akhir)
    c_tot_val.font = font_bold; c_tot_val.number_format = '"Rp "#,##0'
    c_tot_val.alignment = Alignment(horizontal="right", vertical="center"); c_tot_val.border = border_thin
    
    current_row += 2
    ws.row_dimensions[current_row].height = 22
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell_garansi_header = ws.cell(row=current_row, column=1, value=" SYARAT & KETENTUAN GARANSI REPARASI")
    cell_garansi_header.font = font_section; cell_garansi_header.fill = fill_biru_tua; cell_garansi_header.alignment = Alignment(vertical="center")
    
    syarat_teks = [
        "1. Garansi reparasi berlaku selama 30 HARI terhitung sejak tanggal selesai pengerjaan.",
        "2. Garansi HANYA BERLAKU untuk komponen, sparepart, atau jenis kerusakan yang sama pada nota ini.",
        "3. Nota ini dianggap SAH apabila dikirim langsung oleh pihak Manajemen/Teknisi Resmi kami dalam bentuk GAMBAR/PDF.",
        "4. Segala bentuk manipulasi, perubahan tulisan, atau penyalahgunaan nota fisik/digital di luar sistem kami"
    ]
    for teks in syarat_teks:
        current_row += 1
        ws.row_dimensions[current_row].height = 16
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        ws.cell(row=current_row, column=1, value=teks).font = font_garansi
        ws.cell(row=current_row, column=1).alignment = Alignment(vertical="center")
    
    current_row += 2
    ws.row_dimensions[current_row].height = 20
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    ws.cell(row=current_row, column=1, value=f"- {st.session_state.nama_usaha} -").font = Font(name="Arial", size=10, bold=True, color="1E40AF")
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 35; ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 15; ws.column_dimensions['E'].width = 20; ws.column_dimensions['F'].width = 22
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ==================================================
# HALAMAN LOGIN MULTI-TENANT (PENYARING KAMAR UTAMA)
# ==================================================
def halaman_login():
    st.markdown("<h1 style='text-align: center; color: #0E7490;'>❄️ SISTEM OPERASI INTEGRASI SI DINGIN</h1>", unsafe_allow_html=True)
    kol1, col2, kol3 = st.columns([1, 2, 1])
    
    with col2:
        if st.session_state.forgot_mode:
            with st.container(border=True):
                st.subheader("🔑 Pemulihan Akun Owner via Email")
                # Menggunakan logika bawaan email otp lama bos
                st.info("Fitur pemulihan akun master developer pusat.")
                email_input = st.text_input("Masukkan Alamat Email Owner:")
                if st.button("⬅️ Kembali ke Halaman Login"):
                    st.session_state.forgot_mode = False
                    st.rerun()
        else:
            with st.container(border=True):
                username = st.text_input("Username")
                password = st.text_input("Password", type="password")
                
                if st.button("🚪 MASUK KE SISTEM", type="primary", use_container_width=True):
                    # SAKLAR GERBANG 1: AKUN DEV PUSAT (BOS FARHAN)
                    if username == "farhan" and password == "owner123":
                        st.session_state.authenticated = True
                        st.session_state.user_role = "super_admin"
                        st.session_state.user_fullname = "Farhan Kholili (Developer)"
                        st.session_state.username = "farhan"
                        st.session_state.bengkel_id = 1
                        st.rerun()
                    else:
                        # SAKLAR GERBANG 2: USER BENGKEL UMUM / PENYEWA
                        user_match = User.get_or_none(User.username == username, User.password == password)
                        if user_match:
                            # Proteksi: Cek status sewa bengkel, aktif atau diblokir
                            if not user_match.bengkel.status_aktif:
                                st.error("⚠️ Akses Ditangguhkan! Masa langganan aplikasi bengkel Anda telah habis. Silakan hubungi Bos Farhan untuk perpanjangan.")
                            else:
                                st.session_state.authenticated = True
                                st.session_state.user_role = user_match.role
                                st.session_state.user_fullname = user_match.nama_lengkap
                                st.session_state.user_id = user_match.id
                                st.session_state.username = user_match.username
                                
                                # SUNTIK IDENTITAS KAMAR BENGKEL KE MEMORI STREAMLIT
                                st.session_state.bengkel_id = user_match.bengkel.id
                                st.session_state.nama_usaha = user_match.bengkel.nama_bengkel
                                st.session_state.alamat_usaha = user_match.bengkel.alamat_bengkel
                                st.session_state.kontak_usaha = user_match.bengkel.kontak_bengkel
                                st.rerun()
                        else: 
                            st.error("Username atau Password salah, Bos!")

# ==================================================
# INTERFACE TEKNISI LAPANGAN (TERISOLASI)
# ==================================================
def tampilan_teknisi():
    st.sidebar.markdown(f"### 🧑‍🔧 Teknisi: **{st.session_state.user_fullname}**")
    st.sidebar.caption(f"🏢 Perusahaan: {st.session_state.nama_usaha}")
    if st.sidebar.button("🔒 Keluar Aplikasi", use_container_width=True):
        st.session_state.authenticated = False
        st.rerun()
        
    tab_kerja, tab_riwayat_tek = st.tabs(["🛠️ Input Pekerjaan Lapangan", "🖨️ Riwayat & Cetak Ulang Nota"])
    
    with tab_kerja:
        st.title("❄️ INPUT DETAIL NOTA BARU")
        # FILTER: Hanya ambil daftar jasa milik bengkel dia sendiri
        semua_layanan = Layanan.select().where(Layanan.bengkel == st.session_state.bengkel_id)
        kol1, kol2 = st.columns([1, 1])

        with kol1:
            st.subheader("📋 1. Data Pelanggan")
            nama = st.text_input("Nama Pelanggan")
            alamat = st.text_area("Alamat Rumah")
            telp = st.text_input("No. WhatsApp")

            st.divider()
            st.subheader("🛠️ 2. Input Item Pekerjaan")
            opsi = {l.id: f"{l.nama_layanan} (Rp {l.harga:,})" for l in semua_layanan}
            
            if not opsi:
                st.warning("Master data harga layanan belum diatur oleh Owner bengkel Anda.")
                layanan_pilihan = None
            else:
                layanan_pilihan = st.selectbox("Pilih Jenis Pekerjaan", options=list(opsi.keys()), format_func=lambda x: opsi[x])
                
            ruangan = st.text_input("Lokasi / Ruangan (Misal: Kamar Utama)")
            qty = st.number_input("Jumlah (Qty)", min_value=1, value=1, step=1)

            if st.button("➕ Tambah Pekerjaan", use_container_width=True):
                if ruangan == "" or not layanan_pilihan: 
                    st.warning("Data input kerjaan belum lengkap, Bos!")
                else:
                    layanan_obj = [l for l in semua_layanan if l.id == layanan_pilihan][0]
                    st.session_state.keranjang_kerja.append({
                        "id_layanan": layanan_pilihan, "nama_layanan": layanan_obj.nama_layanan,
                        "harga": layanan_obj.harga, "jumlah": qty, "lokasi_ruang": ruangan
                    })
                    st.rerun()

        with kol2:
            st.subheader("🛒 3. Keranjang Kerja")
            if not st.session_state.keranjang_kerja: 
                st.info("Keranjang kosong.")
            else:
                total_nota = 0
                for idx, item in enumerate(st.session_state.keranjang_kerja):
                    subtotal = item["harga"] * item["jumlah"]
                    total_nota += subtotal
                    st.markdown(f"**{idx+1}. {item['nama_layanan']}** ({item['lokasi_ruang']}) x{item['jumlah']}")
                
                st.metric(label="TOTAL NOTA TEMPORER", value=f"Rp {total_nota:,}")
                
                if st.button("🗑️ Reset Keranjang", use_container_width=True):
                    st.session_state.keranjang_kerja = []
                    st.rerun()

                st.divider()
                if st.button("🔥 PROSES & BUAT NOTA INVOICE EXCEL", type="primary", use_container_width=True):
                    if not nama: 
                        st.error("Nama pelanggan wajib diisi!")
                    else:
                        # EKSEKUSI DATA BARU TERSEGMENTASI BENGKEL ID
                        id_tek = st.session_state.user_id
                        order = Order.create(
                            bengkel_id=st.session_state.bengkel_id, # DIKUNCI DI SINI
                            teknisi_id=id_tek,
                            no_invoice=f"INV-{random.randint(10000,99999)}",
                            tanggal_kerja=pd.Timestamp.now().date(),
                            nama_pelanggan=nama,
                            alamat_pelanggan=alamat,
                            no_telp_pelanggan=telp,
                            total_bayar=total_nota,
                            status_pembayaran="Belum Dibayar"
                        )
                        
                        for k in st.session_state.keranjang_kerja:
                            DetailOrder.create(
                                order=order,
                                layanan_id=k["id_layanan"],
                                lokasi_ruang=k["lokasi_ruang"],
                                harga_snapshot=k["harga"],
                                jumlah=k["jumlah"]
                            )
                        
                        buffer_nota = buat_file_excel_nota(order, st.session_state.keranjang_kerja)
                        st.session_state.nota_siap_download = {
                            "label": f"📥 DOWNLOAD NOTA EXCEL ({order.no_invoice})",
                            "data": buffer_nota.getvalue(),
                            "filename": f"Nota_{order.no_invoice}.xlsx"
                        }
                        st.session_state.keranjang_kerja = []
                        st.balloons()
                        st.rerun()
                
                if st.session_state.nota_siap_download is not None:
                    st.download_button(
                        label=st.session_state.nota_siap_download["label"],
                        data=st.session_state.nota_siap_download["data"],
                        file_name=st.session_state.nota_siap_download["filename"],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary", use_container_width=True
                    )

    with tab_riwayat_tek:
        st.subheader(" Riwayat Hasil Inputan Anda")
        query_riwayat_tek = Order.select().where((Order.teknisi_id == st.session_state.user_id) & (Order.bengkel == st.session_state.bengkel_id)).order_by(Order.tanggal_kerja.desc())
        
        if not query_riwayat_tek.exists():
            st.info("Belum ada riwayat nota.")
        else:
            data_tabel_tek = [{"No Invoice": o.no_invoice, "Tanggal": o.tanggal_kerja, "Nama Pelanggan": o.nama_pelanggan, "Status": o.status_pembayaran} for o in query_riwayat_tek]
            st.dataframe(pd.DataFrame(data_tabel_tek), use_container_width=True, hide_index=True)

# ==================================================
# INTERFACE OWNER BENGKEL (TERISOLASI PER USER)
# ==================================================
def tampilan_owner():
    st.sidebar.markdown(f"### 👑 Akses Owner: **{st.session_state.user_fullname}**")
    st.sidebar.caption(f"🏬 Usaha: {st.session_state.nama_usaha}")
    if st.sidebar.button("🔒 Keluar Aplikasi", use_container_width=True):
        st.session_state.authenticated = False
        st.rerun()

    tab_keuangan, tab_layanan, tab_teknisi, tab_toko = st.tabs([
        "📈 Keuangan & Invoice", "⚙️ Kelola Harga Layanan", "🧑‍🔧 Kelola Data Teknisi", "🏨 Pengaturan Toko"
    ])
    
    # --------------------------------------------------
    # TAB KEUANGAN: FIX BUG DUPLICATE FORM KEY PER BENGKEL
    # --------------------------------------------------
    with tab_keuangan:
        # FILTER: Hanya ambil order milik bengkel owner yang sedang login
        query_order = (Order.select(Order, User).join(User, on=(Order.teknisi_id == User.id)).where(Order.bengkel == st.session_state.bengkel_id).order_by(Order.tanggal_kerja.desc()))
        orders_data = []
        for o in query_order:
            orders_data.append({
                "ID": o.id, "No Invoice": o.no_invoice, "Tanggal": o.tanggal_kerja,
                "Pelanggan": o.nama_pelanggan, "Teknisi": o.teknisi.nama_lengkap if o.teknisi else "Tanpa Teknisi",
                "Total Bayar": o.total_bayar, "Status": o.status_pembayaran
            })
        df = pd.DataFrame(orders_data)
        
        if df.empty: 
            st.info("Belum ada transaksi di database bengkel Anda.")
        else:
            t_omset = df[df["Status"] == "Lunas"]["Total Bayar"].sum()
            t_piutang = df[df["Status"] == "Belum Dibayar"]["Total Bayar"].sum()
            
            kol_m1, kol_m2, kol_m3 = st.columns(3)
            kol_m1.metric("💰 TOTAL OMSET MASUK", f"Rp {t_omset:,}")
            kol_m2.metric("⚠️ TOTAL PIUTANG", f"Rp {t_piutang:,}")
            kol_m3.metric("📦 TOTAL NOTA", f"{len(df)} Nota")
            
            st.dataframe(df.drop(columns=["ID"]), use_container_width=True, hide_index=True)
            
            st.divider()
            st.markdown("### 🛠️ PANEL KONTROL & EDIT INVOICE")
            opsi_nota = {row["ID"]: f"{row['No Invoice']} - {row['Pelanggan']} (Rp {row['Total Bayar']:,})" for _, row in df.iterrows()}
            nota_sel = st.selectbox("Pilih Invoice:", options=list(opsi_nota.keys()), format_func=lambda x: opsi_nota[x])
            
            if nota_sel:
                ord_obj = Order.get_by_id(nota_sel)
                query_items = DetailOrder.select().where(DetailOrder.order_id == nota_sel)
                list_items = [{"id_detail": item.id, "nama_layanan": item.layanan.nama_layanan if item.layanan else "Jasa Terhapus", "harga": item.harga_snapshot, "jumlah": item.jumlah, "lokasi_ruang": item.lokasi_ruang} for item in query_items]
                
                kol_btn1, kol_btn2, kol_btn3 = st.columns(3)
                with kol_btn1:
                    if ord_obj.status_pembayaran != "Lunas":
                        if st.button("✅ SET JADI LUNAS", type="primary", use_container_width=True, key=f"lns_{ord_obj.id}"):
                            Order.update(status_pembayaran="Lunas").where(Order.id == ord_obj.id).execute()
                            st.rerun()
                with kol_btn2:
                    if st.button(f"🖨️ CETAK NOTA EXCEL", use_container_width=True, key=f"ctk_{ord_obj.id}"):
                        buffer_cetak = buat_file_excel_nota(ord_obj, list_items)
                        st.session_state.owner_download_file = {"data": buffer_cetak.getvalue(), "filename": f"Nota_{ord_obj.no_invoice}.xlsx"}
                        st.rerun()
                with kol_btn3:
                    if st.button("🚨 HAPUS INVOICE", use_container_width=True, key=f"hps_{ord_obj.id}"):
                        DetailOrder.delete().where(DetailOrder.order_id == ord_obj.id).execute()
                        Order.delete().where(Order.id == ord_obj.id).execute()
                        st.rerun()
                        
                if "owner_download_file" in st.session_state:
                    st.download_button("📥 DOWNLOAD SEKARANG", data=st.session_state.owner_download_file["data"], file_name=st.session_state.owner_download_file["filename"], use_container_width=True, type="primary")
                    if st.button("❌ Selesai"):
                        del st.session_state.owner_download_file
                        st.rerun()

                st.divider()
                # FORM UTAMA EDIT INFORMASI UTAMA INDUK INVOICE (KEY UNIK MULTI-TENANT)
                with st.form(key=f"form_edit_induk_invoice_utama_id_{ord_obj.id}"):
                    st.markdown("##### 📝 Panel Edit Informasi Pelanggan:")
                    edit_nama = st.text_input("Nama Pelanggan:", value=ord_obj.nama_pelanggan)
                    edit_alamat = st.text_area("Alamat:", value=ord_obj.alamat_pelanggan)
                    if st.form_submit_button("🔄 UPDATE DATA NOTA UTAMA", use_container_width=True):
                        Order.update(nama_pelanggan=edit_nama, alamat_pelanggan=edit_alamat).where(Order.id == ord_obj.id).execute()
                        st.success("Data Invoice Utama Berhasil Diperbarui!")
                        st.rerun()

    # --------------------------------------------------
    # TAB LAYANAN: ATUR JASA MILIK BENGKEL MASING-MASING
    # --------------------------------------------------
    with tab_layanan:
        st.subheader("⚙️ Atur Master Jasa & Harga Bengkel")
        with st.form("form_tambah_layanan"):
            n_lay = st.text_input("Nama Layanan Baru:")
            h_lay = st.number_input("Harga Jual (Rp):", min_value=0, step=5000)
            if st.form_submit_button("➕ Tambah Master Layanan"):
                Layanan.create(bengkel_id=st.session_state.bengkel_id, nama_layanan=n_lay, harga=h_lay)
                st.rerun()
                
        layanan_bengkel = Layanan.select().where(Layanan.bengkel == st.session_state.bengkel_id)
        for l in layanan_bengkel:
            st.text(f"• {l.nama_layanan} — Rp {l.harga:,}")

    # --------------------------------------------------
    # TAB TEKNISI: KELOLA KARYAWAN BENGKEL SENDIRI
    # --------------------------------------------------
    with tab_teknisi:
        st.subheader("🧑‍🔧 Karyawan & Teknisi Lapangan Anda")
        with st.form("form_tambah_teknisi"):
            t_nama = st.text_input("Nama Lengkap Teknisi:")
            t_user = st.text_input("Username Login:")
            t_pass = st.text_input("Password Login:")
            if st.form_submit_button("➕ Daftarkan Teknisi"):
                User.create(bengkel_id=st.session_state.bengkel_id, nama_lengkap=t_nama, username=t_user, password=t_pass, role="teknisi")
                st.rerun()

    # --------------------------------------------------
    # TAB TOKO: EDIT PROFIL BENGKEL INDIVIDU
    # --------------------------------------------------
    with tab_toko:
        st.subheader("🏨 Profil Info Nota Bengkel")
        with st.form("form_toko"):
            nama_t = st.text_input("Nama Bengkel Di Nota:", value=st.session_state.nama_usaha)
            alamat_t = st.text_area("Slogan/Alamat Di Nota:", value=st.session_state.alamat_usaha)
            kontak_t = st.text_input("No Telp Bengkel:", value=st.session_state.kontak_usaha)
            if st.form_submit_button("💾 Simpan Pengaturan Toko"):
                Bengkel.update(nama_bengkel=nama_t, alamat_bengkel=alamat_t, kontak_bengkel=kontak_t).where(Bengkel.id == st.session_state.bengkel_id).execute()
                st.session_state.nama_usaha = nama_t
                st.session_state.alamat_usaha = alamat_t
                st.session_state.kontak_usaha = kontak_t
                st.success("Profil nota usaha diperbarui!")
                st.rerun()

# ==================================================
# 👑 SAKLAR KHUSUS: HALAMAN SUPER ADMIN (BOS FARHAN)
# ==================================================
def tampilan_super_admin():
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 👑 PUSAT DEVELOPER")
    if st.sidebar.button("🔒 Keluar Super Admin", use_container_width=True):
        st.session_state.authenticated = False
        st.rerun()
        
    menu_dev = st.sidebar.radio("Navigasi Kontrol:", ["🏢 Kelola Bengkel / Tenant", "📊 Pantau Semua Transaksi Global"])
    
    if menu_dev == "🏢 Kelola Bengkel / Tenant":
        st.title("🏢 Panel Manajemen Bengkel Langganan")
        st.caption("Pusat saklar hidup-mati akun aplikasi penyewa SaaS milik Bos Farhan.")
        
        with st.expander("➕ Daftarkan Bengkel / Client Baru", expanded=True):
            with st.form("form_tambah_bengkel_baru", clear_on_submit=True):
                n_bengkel = st.text_input("Nama Bengkel Baru:")
                a_bengkel = st.text_area("Alamat Lengkap Bengkel:")
                k_bengkel = st.text_input("Kontak / No. WA Owner Bengkel:")
                st.markdown("##### 🔑 Buat Akun Owner Pertama Bengkel Tersebut:")
                u_owner = st.text_input("Username Owner (Harus Unik):")
                p_owner = st.text_input("Password Owner:", type="password")
                
                if st.form_submit_button("🚀 AKTIFKAN & BUAT AKUN CLIENT", type="primary", use_container_width=True):
                    if not n_bengkel or not u_owner or not p_owner:
                        st.error("Gagal! Field pendaftaran tidak boleh kosong, Bos!")
                    else:
                        cek_user = User.get_or_none(User.username == u_owner)
                        if cek_user:
                            st.error("⚠️ Username tersebut sudah dipakai bengkel lain, cari nama lain Bos!")
                        else:
                            b_baru = Bengkel.create(nama_bengkel=n_bengkel, alamat_bengkel=a_bengkel, kontak_bengkel=k_bengkel, status_aktif=True)
                            User.create(bengkel=b_baru, nama_lengkap=f"Owner {n_bengkel}", username=u_owner, password=p_owner, role="owner")
                            st.success(f"✨ Sukses! Bengkel '{n_bengkel}' aktif. Username owner: {u_owner}")
                            st.rerun()

        st.divider()
        st.markdown("#### 📋 Daftar Seluruh Bengkel Aktif & Status Sewa")
        
        semua_bengkel = Bengkel.select()
        for b in semua_bengkel:
            if b.id == 1: continue # Jangan kunci bengkel pusat master
            
            lbl_status = "🟢 AKTIF / LANCAR" if b.status_aktif else "🔴 NONAKTIF / TERKUNCI"
            with st.container(border=True):
                k1, k2 = st.columns([3, 1])
                with k1:
                    st.markdown(f"### {b.nama_bengkel.upper()}")
                    st.text(f"📍 Alamat: {b.alamat_bengkel} | 📞 Kontak: {b.kontak_bengkel}")
                    st.markdown(f"**Kondisi Finansial Sewa:** `{lbl_status}`")
                with k2:
                    st.markdown("<div style='padding-top:20px;'></div>", unsafe_allow_html=True)
                    if b.status_aktif:
                        if st.button("🔒 MATIKAN AKSES", key=f"blk_{b.id}", type="primary", use_container_width=True):
                            Bengkel.update(status_aktif=False).where(Bengkel.id == b.id).execute()
                            st.rerun()
                    else:
                        if st.button("🔓 AKTIFKAN AKSES", key=f"opn_{b.id}", use_container_width=True):
                            Bengkel.update(status_aktif=True).where(Bengkel.id == b.id).execute()
                            st.rerun()

    elif menu_dev == "📊 Pantau Semua Transaksi Global":
        st.title("📊 Data Omset Global Seluruh Client")
        query_global = (Order.select(Order, Bengkel).join(Bengkel).order_by(Order.tanggal_kerja.desc()))
        data_g = [{"Nama Bengkel": o.bengkel.nama_bengkel, "No Invoice": o.no_invoice, "Tanggal": o.tanggal_kerja, "Pelanggan": o.nama_pelanggan, "Total": o.total_bayar, "Status": o.status_pembayaran} for o in query_global]
        
        if not data_g:
            st.info("Belum ada data transaksi masuk di server pusat.")
        else:
            df_g = pd.DataFrame(data_g)
            st.dataframe(df_g, use_container_width=True, hide_index=True)
            t_duit = df_g[df_g["Status"] == "Lunas"]["Total"].sum()
            st.metric("💰 TOTAL DUIT LUNAS GLOBAL (ALL CLIENT)", f"Rp {t_duit:,}")

# ==================================================
# ALUR NAVIGASI UTAMA ROUTER RUNNER
# ==================================================
if not st.session_state.authenticated:
    halaman_login()
else:
    if st.session_state.user_role == "super_admin":
        tampilan_super_admin()
    elif st.session_state.user_role == "owner":
        tampilan_owner()
    elif st.session_state.user_role == "teknisi":
        tampilan_teknisi()
